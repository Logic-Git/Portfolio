"""
Email Generation module for the automated cold emailing system.
Provides functions for generating personalized cold emails for leads.
"""

import logging
from openpyxl import load_workbook
import google.generativeai as genai
from utils import rate_limit
from research import research_company_with_gemini, extract_company_info_for_campaign


@rate_limit(calls_per_minute=10)
def generate_email_content(prompt):
    """
    Generates email content using the Gemini API.

    Args:
        prompt: The prompt to send to the model.

    Returns:
        The generated email content, or None if an error occurred.
    """
    model = genai.GenerativeModel("gemini-1.5-flash")
    try:
        response = model.generate_content(prompt)
        return response.text
    except Exception as e:
        logging.error(f"Error during Gemini API request: {e}")
        return None


@rate_limit(calls_per_minute=10)
def generate_emails_for_leads(
    lead_directory_file,
    template_email,
    template_email_sequence,
    row_start,
    row_end,
    sender_name,
):
    """
    Generates tailored cold emails for a range of leads in the lead directory using the Gemini API.

    Args:
        lead_directory_file (str): Path to the Excel file containing the lead directory.
        template_email (str): The template for the initial email.
        template_email_sequence (list): List of templates for follow-up emails.
        row_start (int): The starting row number (inclusive) for processing leads.
        row_end (int): The ending row number (inclusive) for processing leads.
        sender_name (str): The name of the sender to be used in the emails.

    Returns:
        bool: True if all emails were generated successfully, False otherwise.
    """
    try:
        workbook = load_workbook(lead_directory_file)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]

        # Check if the Excel file has the required columns
        required_email_columns = [
            f"Email {i + 1}" for i in range(len(template_email_sequence) + 1)
        ]
        required_sent_columns = [
            f"Email {i + 1} Sent" for i in range(len(template_email_sequence) + 1)
        ]

        missing_email_columns = [
            col for col in required_email_columns if col not in headers
        ]
        missing_sent_columns = [
            col for col in required_sent_columns if col not in headers
        ]

        if missing_email_columns:
            logging.error(
                f"Error: Missing email columns in Excel file: {missing_email_columns}"
            )
            return False
        if missing_sent_columns:
            logging.error(
                f"Error: Missing 'sent' columns in Excel file: {missing_sent_columns}"
            )
            return False

        for row_index in range(row_start, row_end + 1):
            lead_info = {
                headers[i]: sheet.cell(row=row_index, column=i + 1).value
                for i in range(len(headers))
            }
            company_name = lead_info.get(
                "Company Name"
            )  # Assuming "Company Name" is the header
            lead_email = lead_info.get("Email")  # Assuming "Email" is the header

            if not company_name or not lead_email:
                logging.warning(
                    f"Skipping row {row_index}: Company name or email missing."
                )
                continue

            # Extract company info from the template email
            company_info = extract_company_info_for_campaign(
                template_email, template_email_sequence
            )

            # Research the company using Gemini
            research_summary = research_company_with_gemini(company_name, company_info)

            if research_summary is None:
                logging.error(f"Research failed for {company_name} (row {row_index}).")
                continue

            # Generate initial email
            initial_email_prompt = f"""
                You are a highly skilled copywriter specializing in crafting compelling cold emails for a company.
                You are also the best sales person in the world capable of convincing anyone to buy anything, and you can do this while staying ethical and not lying.
                Now, I want you to use your skills to write cold emails for our company. 
                Use your sales skills to make the email compelling, convincing the recipient to take some action like book a meeting, try a demo etc.
                Remember that convincing the recipient to buy what we are selling is not important for you. 
                The most importnat thing for you is to identify which action/actions would be easiest to convince them on taking and then convince them on taking that.
                So even if you can't convince the recipient to buy what we are selling, but are able to convince them to take a meeting with us, you get maximum possible credit.
                I want you to work very hard on every cold email, ensuring a perfect piece is crefted at the end of every response you give.
                Here is a template email that we have used successfully in the past:\\n```\\n{template_email}\\n```\\n\n
                Here is some information about the company we are targeting:\\n```\\nCompany Name: {company_name}\\nResearch Summary:\\n{research_summary}\\n```\\n\n
                You can use the template to understand what we are offering, what is the value of what we are offering to the customeer and how we market it.
                Also use the template to learn the type of language we use in our cold emails. Use the same language.
                Use target company information to tailor the email to the person we are emailing.
                Here, you should aim to make the person feel as if we have spent a lot of time researching specifically on them, 
                and that they are a great fit for our service.
                Also use the target company information to emphasize more on benefits that we are offering that would be more relevant to the target company.
                The email should:
                - Be in HTML format ready to be sent directly, without any extra comments or formatting. An example of what I would like the output to look like is the following:
                ```
                <html> 
                <body> 
                <p>This is an <b>HTML</b> email sent from Python using the Gmail API.</p> 
                <p>An interesting thing about this email is that it is sent fully programatically</p>
                <p>All the best</p> 
                </body> 
                </html>
                ```
                Note: The above is not indicative in any way of the type of content I want in your output. 
                It is just there to show you how you should write the HTML code and output it after you have decided on the content from my other instructions.
                Therefore, no need to do anything additional like enclosing the output in backticks, adding things in <title> tags etc.
                - Be very direct in the email. Only write words that have a purpose. 
                Using correct Grammar is not important. Your aim should be to get the message across without boring the recipient or being unnecessarily long.
                Example sentences are:
                ```
                Hey [First Name] – great to connect with you, and congrats on all the success at [Company Name]!
                And here's the best part: our model works directly with your employees' insurance, ensuring no cost to your company while enhancing productivity.
                Care to carve out 15 minutes to chat?
                Hey [First Name], maybe third time's a charm...?
                ```
                The above sentences might not be correct gramatically but anyone reading them can understand what they mean, and they are concise so they do not bore the reader.
                Also, this is how humans write so the reader reading the email would feel that there is a human writing the email.
                You should also do this. 
                So to sum up, when writing your email, your style should be similar to the sentences given above
                and when reading the email, the reader should feel as if a human wrote the email.
                - In the email, stay hyper focused on the value that the recipient will get from the service instead of using jargon, explaining functionality or other irrelevant things.
                For example:
                ```
                Suppose we provide an invoice generating software to other companies.
                Instead of saying: 
                Software uses LLM that learns from previously generated invoices to generate new invoices and invoice templates quickly
                You should say:
                Automatically generate invoices, saving 60% time on invoice generation.
                Save at least 30% on software costs too
                ```
                - Add as many metrics as possible if you have knowledge of them from things I have given you
                - Address the recipient by their first name (which will be given below), whenever needed.
                - Use "{sender_name}" as the sender's name at the end of the email and wherever else in the email you see its usage fit
                Please generate a tailored cold email for {company_name}, addressed to {lead_info.get("First Name", "")}.
                """
            initial_email = generate_email_content(initial_email_prompt)

            if initial_email:
                # Save the initial email and mark it as not sent
                sheet.cell(
                    row=row_index,
                    column=headers.index("Email 1") + 1,
                    value=initial_email,
                )
                sheet.cell(
                    row=row_index, column=headers.index("Email 1 Sent") + 1, value=0
                )

                # Generate follow-up emails
                current_email_chain = [initial_email]
                for i, follow_up_template in enumerate(template_email_sequence):
                    follow_up_prompt = f"""
                        You are a highly skilled copywriter specializing in crafting compelling cold emails for a company.
                        You are also the best sales person in the world capable of convincing anyone to buy anything, and you can do this while staying ethical and not lying.
                        Now, I want you to use your skills to write cold emails for our company. 
                        Use your sales skills to make the email compelling, convincing the recipient to take some action like book a meeting, try a demo etc.
                        Remember that convincing the recipient to buy what we are selling is not important for you. 
                        The most importnat thing for you is to identify which action/actions would be easiest to convince them on taking and then convince them on taking that.
                        So even if you can't convince the recipient to buy what we are selling, but are able to convince them to take a meeting with us, you get maximum possible credit.
                        I want you to work very hard on every cold email, ensuring a perfect piece is crefted at the end of every response you give.
                        Here is a sequence of email templates our company has used in the past:\\n```\\nInitial Email:\\n{template_email}\\n\n
                        {" " if i == 0 else f"Follow-up {i}:"}\\n{" " if i == 0 else follow_up_template}\\n```\\n\n
                        Here is the email chain we have sent to {company_name} so far:\\n```\\n{chr(10).join(f"Email {j + 1}:{email}" for j, email in enumerate(current_email_chain))}\\n```\\n\n
                        And here is some information about the company we are targeting:\\n```\\nCompany Name: {company_name}\\nResearch Summary:\\n{research_summary}\\n```\\n\n
                        You can use the templates to understand what we are offering, what is the value of what we are offering to the customeer and how we market it.
                        Also use the template to learn the type of language we use in our cold emails. Use the same language.
                        Use target company information to tailor the email to the person we are emailing.
                        Here, you should aim to make the person feel as if we have spent a lot of time researching specifically on them, 
                        and that they are a great fit for our service.
                        Also use the target company information to emphasize more on benefits that we are offering that would be more relevant to the target company.
                        Your task is to write Email {i + 2} (the next email) in the sequence. 
                        Use the past emails we have sent to this company to maintain context
                        The email should:
                        - Be in HTML format ready to be sent directly, without any extra comments or formatting. An example of what I would like the output to look like is the following:
                        ```
                        <html> 
                        <body> 
                        <p>This is an <b>HTML</b> email sent from Python using the Gmail API.</p> 
                        <p>An interesting thing about this email is that it is sent fully programatically</p>
                        <p>All the best</p> 
                        </body> 
                        </html>
                        ```
                        Note: The above is not indicative in any way of the type of content I want in your output. 
                        It is just there to show you how you should write the HTML code and output it after you have decided on the content from my other instructions.
                        Therefore, no need to do anything additional like enclosing the output in backticks, adding things in <title> tags etc.
                        - Be very direct in the email. Only write words that have a purpose. 
                        Using correct Grammar is not important. Your aim should be to get the message across without boring the recipient or being unnecessarily long.
                        Example sentences are:
                        ```
                        Hey [First Name] – great to connect with you, and congrats on all the success at [Company Name]!
                        And here's the best part: our model works directly with your employees' insurance, ensuring no cost to your company while enhancing productivity.
                        Care to carve out 15 minutes to chat?
                        Hey [First Name], maybe third time's a charm...?
                        ```
                        The above sentences might not be correct gramatically but anyone reading them can understand what they mean, and they are concise so they do not bore the reader.
                        Also, this is how humans write so the reader reading the email would feel that there is a human writing the email.
                        You should also do this. 
                        So to sum up, when writing your email, your style should be similar to the sentences given above
                        and when reading the email, the reader should feel as if a human wrote the email.
                        - In the email, stay hyper focused on the value that the recipient will get from the service instead of using jargon, explaining functionality or other irrelevant things.
                        For example:
                        ```
                        Suppose we provide an invoice generating software to other companies.
                        Instead of saying: 
                        Software uses LLM that learns from previously generated invoices to generate new invoices and invoice templates quickly
                        You should say:
                        Automatically generate invoices, saving 60% time on invoice generation.
                        Save at least 30% on software costs too
                        ```
                        - Add as many metrics as possible if you have knowledge of them from things I have given you
                        - Also remember that as the chain moves along, slowly the emails should get more and more brief. 
                        For example Email 5 should be much shorter than Email 3. This is exemplified by the template chain.
                        This is because as more and more follow-ups are sent, we do not want the recipient to feel that we are bugging or disturbing them.
                        At the same time, we want to remind them of what we are offering because there is a chance that they find what we are offering useful but got busy so couldn't reply to us.
                        - Address the recipient by their first name (which will be given below), whenever needed.
                        - Use "{sender_name}" as the sender's name at the end of the email and wherever else in the email you see its usage fit.
                        Please generate follow-up email number {i + 2} for {company_name}, addressed to {lead_info.get("First Name", "")}.
                        """
                    follow_up_email = generate_email_content(follow_up_prompt)
                    if follow_up_email:
                        # Save the follow-up email and mark it as not sent
                        sheet.cell(
                            row=row_index,
                            column=headers.index(f"Email {i + 2}") + 1,
                            value=follow_up_email,
                        )
                        sheet.cell(
                            row=row_index,
                            column=headers.index(f"Email {i + 2} Sent") + 1,
                            value=0,
                        )
                        current_email_chain.append(follow_up_email)
                    else:
                        # If any email in the sequence fails to generate, don't save any data for this lead
                        logging.warning(
                            f"Failed to generate follow-up email {i + 2} for {company_name} (row {row_index}). Skipping this lead."
                        )
                        for j in range(
                            1, i + 2
                        ):  # Clear any previously saved emails for this lead
                            sheet.cell(
                                row=row_index,
                                column=headers.index(f"Email {j}") + 1,
                                value=None,
                            )
                            sheet.cell(
                                row=row_index,
                                column=headers.index(f"Email {j} Sent") + 1,
                                value=None,
                            )
                        break  # Move on to the next lead

            # Save the workbook after processing each lead
            workbook.save(lead_directory_file)

        return True

    except FileNotFoundError:
        logging.error(f"Error: Lead directory file '{lead_directory_file}' not found.")
        return False
    except Exception as e:
        logging.error(f"An unexpected error occurred: {e}")
        return False
