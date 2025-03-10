"""
Email Sending module for the automated cold emailing system.
Provides functions for sending cold emails via Gmail API.
"""

import base64
import os
import pickle
from email.mime.text import MIMEText
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from openpyxl import load_workbook
import logging

# Gmail API scope
SCOPES = [
    "https://www.googleapis.com/auth/gmail.send",
    "https://www.googleapis.com/auth/gmail.modify",
]

# File to store email-token mapping
EMAIL_TO_TOKEN_FILE = "email_to_token.pickle"

# Excel file for email-credentials mapping
EMAIL_TO_CREDENTIALSJSON = "Email-to-Credentials.xlsx"


def send_email(sender, to, subject, message_text, sender_name=None, thread_id=None):
    """
    Sends an email using the Gmail API.

    Args:
        sender: Email address of the sender.
        to: Email address of the receiver.
        subject: The subject of the email message.
        message_text: The text of the email message.
        sender_name (str, optional): The name of the sender. If None, only email will be used.
        thread_id (str, optional): The ID of the thread to reply to. If None, a new thread is created.

    Returns:
        The thread ID of the sent email.
    """

    def create_message(sender, to, subject, message_text, sender_name=None):
        """Create a message for an email."""
        message = MIMEText(message_text, "html")
        message["to"] = to

        if sender_name:
            message["from"] = f"{sender_name} <{sender}>"  # Include sender name
        else:
            message["from"] = sender

        message["subject"] = subject
        return {"raw": base64.urlsafe_b64encode(message.as_string().encode()).decode()}

    def send_message(service, user_id, message, thread_id=None):
        """Send an email message."""
        try:
            if thread_id:
                # Get the original message
                original_message = (
                    service.users()
                    .messages()
                    .get(userId=user_id, id=thread_id, format="metadata")
                    .execute()
                )  # Fetch only metadata

                # Extract the threadId and References header
                thread_id = original_message["threadId"]
                references = original_message["payload"]["headers"]

                # Find the References header
                references_header = None
                for header in references:
                    if header["name"] == "References":
                        references_header = header["value"]
                        break

                # Set the In-Reply-To and References headers
                message["In-Reply-To"] = original_message["id"]
                message["References"] = (
                    f"{references_header} {original_message['id']}"
                    if references_header
                    else original_message["id"]
                )

                # Send the message as a reply
                message = (
                    service.users()
                    .messages()
                    .send(userId=user_id, body=message, threadId=thread_id)
                    .execute()
                )
                print("Message Id: %s" % message["id"])
                return message
            else:
                message = (
                    service.users()
                    .messages()
                    .send(userId=user_id, body=message)
                    .execute()
                )
                print("Message Id: %s" % message["id"])
                print("Thread Id: %s" % message["threadId"])  # Print the threadId
                return message["threadId"]  # Return the threadId for new messages
        except Exception as error:
            print("An error occurred: %s" % error)

    def get_credentials(sender):
        """Retrieves the credentials file name for the given sender from the Excel file."""
        try:
            workbook = load_workbook(EMAIL_TO_CREDENTIALSJSON)
            sheet = workbook["Sheet1"]  # Assuming the mapping is in Sheet1

            for row in sheet.iter_rows(
                min_row=2, values_only=True
            ):  # Start from the second row
                if row[0] == sender:  # Assuming email is in the first column
                    return row[
                        1
                    ]  # Return the credentials file name from the second column

            print(f"Error: No credentials found for sender: {sender}")
            exit(1)

        except FileNotFoundError:
            print(f"Error: Credentials file not found: {EMAIL_TO_CREDENTIALSJSON}")
            exit(1)

    # Load or create the email-to-token mapping
    if os.path.exists(EMAIL_TO_TOKEN_FILE):
        with open(EMAIL_TO_TOKEN_FILE, "rb") as f:
            email_to_token = pickle.load(f)
    else:
        email_to_token = {}

    # Get the token file name for the sender
    token_file = email_to_token.get(sender)

    creds = None
    if token_file and os.path.exists(token_file):
        with open(token_file, "rb") as token:
            creds = pickle.load(token)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            credentials_json = get_credentials(sender)
            flow = InstalledAppFlow.from_client_secrets_file(credentials_json, SCOPES)
            creds = flow.run_local_server(port=0)

        # Generate a new token file name if it doesn't exist
        if not token_file:
            token_file = f"token_{len(email_to_token) + 1}.pickle"
            email_to_token[sender] = token_file  # Update the mapping

        # Save the credentials to the token file
        with open(token_file, "wb") as token:
            pickle.dump(creds, token)

    # Save the updated email-to-token mapping
    with open(EMAIL_TO_TOKEN_FILE, "wb") as f:
        pickle.dump(email_to_token, f)

    service = build("gmail", "v1", credentials=creds)

    # Send the email
    message = create_message(sender, to, subject, message_text, sender_name)
    return send_message(
        service, "me", message, thread_id
    )  # Pass thread_id to send_message


def execute_cold_email_campaign(
    lead_directory_file,
    sheet_name,
    email_column_names,
    sending_emails,
    sender_name,
    email_limit,
    analytics_file=None,
    analytics_sheet=None,
    analytics_cell=None,
):
    """
    Executes a cold email campaign by sending emails to leads in a lead directory.

    Args:
        lead_directory_file (str): Path to the Excel file containing the lead directory.
        sheet_name (str): The name of the sheet containing the leads.
        email_column_names (list): List of column names containing the emails to be sent.
        sending_emails (list): List of sender email addresses to use.
        sender_name (str): The name of the sender to be used in the emails.
        email_limit (int): The maximum number of emails to send from each email address.
        analytics_file (str, optional): Path to the Excel file for storing analytics.
        analytics_sheet (str, optional): The name of the sheet for storing analytics.
        analytics_cell (str, optional): The cell coordinate (e.g., "A1") for storing the total sent emails count.
    """
    try:
        workbook = load_workbook(lead_directory_file)
        sheet = workbook[sheet_name]
        headers = [cell.value for cell in sheet[1]]

        # Check if the Excel file has the required columns
        required_sent_columns = [f"{col_name} Sent" for col_name in email_column_names]
        missing_sent_columns = [
            col for col in required_sent_columns if col not in headers
        ]

        if missing_sent_columns:
            logging.error(
                f"Error: Missing 'sent' columns in Excel file: {missing_sent_columns}"
            )
            return False

        email_count = 0
        email_index = 0
        current_email = sending_emails[email_index]

        for col_index, column_name in enumerate(email_column_names):
            if column_name not in headers:
                logging.error(
                    f"Error: Column '{column_name}' not found in the Excel file."
                )
                continue

            sent_column_name = f"{column_name} Sent"
            sent_col_index = headers.index(sent_column_name) + 1

            for row_index in range(
                2, sheet.max_row + 1
            ):  # Start from the second row (index 1)
                if email_count >= email_limit * len(sending_emails):
                    print("Email limit reached for all email addresses.")
                    break

                sent_cell = sheet.cell(row=row_index, column=sent_col_index)
                if sent_cell.value == 1:
                    continue  # Skip if email already sent

                email_cell = sheet.cell(
                    row=row_index, column=headers.index(column_name) + 1
                )
                if email_cell.value is None:
                    print("All emails sent from the specified column.")
                    break  # Exit the inner loop if no more emails in the column

                # Extract lead's first name
                lead_first_name = sheet.cell(
                    row=row_index, column=headers.index("First Name") + 1
                ).value

                # Generate email subject
                subject = f"{sender_name.split()[0]} and {lead_first_name}"

                # Send the email
                send_email(
                    sender=current_email,
                    to=sheet.cell(
                        row=row_index, column=headers.index("Email") + 1
                    ).value,
                    subject=subject,
                    message_text=email_cell.value,
                    sender_name=sender_name,
                )

                # Mark the email as sent
                sent_cell.value = 1

                email_count += 1

                # Switch to the next email address if the limit is reached
                if email_count % email_limit == 0:
                    email_index = (email_index + 1) % len(sending_emails)
                    current_email = sending_emails[email_index]

                # Save the workbook after sending each email
                workbook.save(lead_directory_file)

        # Update analytics if provided
        if analytics_file and analytics_sheet and analytics_cell:
            try:
                analytics_workbook = load_workbook(analytics_file)
                analytics_sheet_obj = analytics_workbook[analytics_sheet]
                analytics_cell_obj = analytics_sheet_obj[analytics_cell]
                analytics_cell_obj.value = (
                    int(analytics_cell_obj.value or 0) + email_count
                )
                analytics_workbook.save(analytics_file)
            except Exception as e:
                logging.error(f"Error updating analytics: {e}")

        return True

    except FileNotFoundError:
        print(f"Error: Lead directory file '{lead_directory_file}' not found.")
        return False
    except KeyError:
        print(f"Error: Sheet '{sheet_name}' not found in the Excel file.")
        return False
    except Exception as e:
        print(f"An error occurred: {e}")
        return False
