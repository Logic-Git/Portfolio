"""
Main module for the automated cold emailing system.
Provides a command-line interface to run the system.
"""

import argparse
from utils import (
    preprocess_csv, extract_lead_info_from_csv, eliminate_incomplete_leads,
    clean_emails, add_opt_out_message, email_verifier, transfer_data_between_files
)
from email_generation import generate_emails_for_leads
from email_sending import execute_cold_email_campaign
from dotenv import load_dotenv
import logging

# Configure logging
LOG_FILE = "email_automation.log"
logging.basicConfig(
    filename=LOG_FILE, 
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

def setup_parser():
    """Set up the command line argument parser."""
    parser = argparse.ArgumentParser(description='Automated Cold Emailing System')
    subparsers = parser.add_subparsers(dest='command', help='Command to run')
    
    # Preprocess CSV
    preprocess_parser = subparsers.add_parser('preprocess', help='Preprocess a CSV file')
    preprocess_parser.add_argument('csv_file', help='Path to the CSV file')
    preprocess_parser.add_argument('columns', nargs='+', help='Column names to preprocess')
    
    # Extract lead info
    extract_parser = subparsers.add_parser('extract', help='Extract lead info from CSV to Excel')
    extract_parser.add_argument('csv_file', help='Path to the CSV file')
    extract_parser.add_argument('excel_file', help='Path to the Excel file')
    extract_parser.add_argument('--mapping', help='JSON string of column mapping')
    
    # Eliminate incomplete leads
    eliminate_parser = subparsers.add_parser('eliminate', help='Eliminate incomplete leads')
    eliminate_parser.add_argument('excel_file', help='Path to the Excel file')
    eliminate_parser.add_argument('sheet_name', help='Sheet name')
    eliminate_parser.add_argument('columns', nargs='+', help='Columns to check')
    
    # Generate emails
    generate_parser = subparsers.add_parser('generate', help='Generate emails for leads')
    generate_parser.add_argument('excel_file', help='Path to the Excel file')
    generate_parser.add_argument('template_file', help='Path to the template file')
    generate_parser.add_argument('--row_start', type=int, default=2, help='Starting row (default: 2)')
    generate_parser.add_argument('--row_end', type=int, help='Ending row')
    generate_parser.add_argument('--sender_name', required=True, help='Sender name')
    
    # Clean emails
    clean_parser = subparsers.add_parser('clean', help='Clean email content')
    clean_parser.add_argument('excel_file', help='Path to the Excel file')
    clean_parser.add_argument('sheet_name', help='Sheet name')
    clean_parser.add_argument('columns', nargs='+', help='Email columns to clean')
    clean_parser.add_argument('--row_start', type=int, default=2, help='Starting row (default: 2)')
    clean_parser.add_argument('--row_end', type=int, required=True, help='Ending row')
    
    # Add opt-out message
    optout_parser = subparsers.add_parser('optout', help='Add opt-out message')
    optout_parser.add_argument('excel_file', help='Path to the Excel file')
    optout_parser.add_argument('sheet_name', help='Sheet name')
    optout_parser.add_argument('columns', nargs='+', help='Email columns to update')
    optout_parser.add_argument('--row_start', type=int, default=2, help='Starting row (default: 2)')
    optout_parser.add_argument('--row_end', type=int, required=True, help='Ending row')
    
    # Verify emails
    verify_parser = subparsers.add_parser('verify', help='Verify emails')
    verify_parser.add_argument('csv_file', help='Path to the CSV file with verified emails')
    verify_parser.add_argument('excel_file', help='Path to the Excel file to verify')
    verify_parser.add_argument('--output', help='Path to save output file')
    verify_parser.add_argument('--excel_col', default='Email', help='Excel email column')
    verify_parser.add_argument('--csv_col', default='email', help='CSV email column')
    
    # Send emails
    send_parser = subparsers.add_parser('send', help='Send cold emails')
    send_parser.add_argument('excel_file', help='Path to the Excel file')
    send_parser.add_argument('sheet_name', default='Sheet1', help='Sheet name')
    send_parser.add_argument('email_columns', nargs='+', help='Email columns to send')
    send_parser.add_argument('sending_emails', nargs='+', help='Sender email addresses')
    send_parser.add_argument('--sender_name', required=True, help='Sender name')
    send_parser.add_argument('--email_limit', type=int, default=50, help='Email limit per sender')
    send_parser.add_argument('--analytics_file', help='Analytics file path')
    send_parser.add_argument('--analytics_sheet', help='Analytics sheet name')
    send_parser.add_argument('--analytics_cell', help='Analytics cell (e.g., A1)')
    
    return parser

def load_templates(template_file):
    """Load email templates from a file."""
    try:
        with open(template_file, 'r') as f:
            content = f.read()
            
        sections = content.split('---')
        if len(sections) < 2:
            raise ValueError("Template file must contain initial email and at least one follow-up separated by '---'")
            
        initial_email = sections[0].strip()
        follow_ups = [section.strip() for section in sections[1:]]
        
        return initial_email, follow_ups
    except Exception as e:
        logging.error(f"Error loading templates: {e}")
        return None, None

def main():
    """Main function to run the automated cold emailing system."""
    # Load environment variables
    load_dotenv()
    
    # Set up command line parser
    parser = setup_parser()
    args = parser.parse_args()
    
    # Process commands
    if args.command == 'preprocess':
        preprocess_csv(args.csv_file, args.columns)
    
    elif args.command == 'extract':
        import json
        mapping = json.loads(args.mapping) if args.mapping else {
            "First Name": "First Name", 
            "Last Name": "Last Name", 
            "Email": ["Recommended Email", "Recommended Work Email", "Recommended Personal Email"],
            "Company Name": "Employer"
        }
        extract_lead_info_from_csv(args.csv_file, args.excel_file, mapping)
    
    elif args.command == 'eliminate':
        eliminate_incomplete_leads(args.excel_file, args.sheet_name, args.columns)
    
    elif args.command == 'generate':
        initial_email, follow_ups = load_templates(args.template_file)
        if not initial_email or not follow_ups:
            return
        
        row_end = args.row_end if args.row_end else None
        if not row_end:
            # Try to determine the maximum row from the Excel file
            from openpyxl import load_workbook
            try:
                wb = load_workbook(args.excel_file)
                sheet = wb.active
                row_end = sheet.max_row
            except Exception as e:
                logging.error(f"Could not determine max row: {e}")
                return
        
        generate_emails_for_leads(
            args.excel_file, initial_email, follow_ups, args.row_start, row_end, args.sender_name
        )
    
    elif args.command == 'clean':
        clean_emails(args.excel_file, args.sheet_name, args.columns, args.row_start, args.row_end)
    
    elif args.command == 'optout':
        add_opt_out_message(args.excel_file, args.sheet_name, args.columns, args.row_start, args.row_end)
    
    elif args.command == 'verify':
        output_file = args.output if args.output else args.excel_file
        mapping = {
            "excel_email_column": args.excel_col, 
            "csv_email_column": args.csv_col
        }
        email_verifier(args.csv_file, args.excel_file, mapping, output_file)
    
    elif args.command == 'send':
        analytics_params = {}
        if all([args.analytics_file, args.analytics_sheet, args.analytics_cell]):
            analytics_params = {
                'analytics_file': args.analytics_file,
                'analytics_sheet': args.analytics_sheet,
                'analytics_cell': args.analytics_cell
            }
            
        execute_cold_email_campaign(
            args.excel_file, args.sheet_name, args.email_columns,
            args.sending_emails, args.sender_name, args.email_limit,
            **analytics_params
        )
    
    else:
        parser.print_help()

if __name__ == "__main__":
    main()