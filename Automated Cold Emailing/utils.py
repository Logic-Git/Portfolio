"""
Utility functions for the automated cold emailing system.
Contains rate limiting and other helper functions.
"""

import time
import logging
import csv
import pandas as pd
from openpyxl import load_workbook, Workbook

# Configure logging
LOG_FILE = "email_generation.log"
logging.basicConfig(
    filename=LOG_FILE,
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
)


# --- Rate Limiting Decorator ---
def rate_limit(calls_per_minute):
    """
    Decorator factory to create rate-limiting decorators with different limits.

    Args:
        calls_per_minute: The maximum number of calls allowed per minute.

    Returns:
        A rate-limiting decorator function.
    """

    def decorator(func):
        """
        Decorator to rate-limit a function.
        """
        call_times = []

        def wrapper(*args, **kwargs):
            """
            Wrapper function that performs rate-limiting.
            """
            now = time.time()
            call_times.append(now)

            # Remove calls older than a minute
            call_times[:] = [t for t in call_times if t > now - 60]

            if len(call_times) > calls_per_minute:
                sleep_time = 60 - (now - call_times[0])
                logging.info(
                    f"Rate limit hit for {func.__name__}. Sleeping for {sleep_time:.2f} seconds."
                )
                time.sleep(sleep_time)
            return func(*args, **kwargs)

        return wrapper

    return decorator


def add_lead_to_excel(excel_filename, lead_email, lead_data):
    """
    Adds lead data and generated emails to an Excel file.

    Args:
        excel_filename: The name of the Excel file.
        lead_email: The email address of the lead to find.
        lead_data: A dictionary containing the lead information and generated emails.

    Returns:
        True if the lead data was added successfully, False otherwise.
    """
    try:
        workbook = load_workbook(excel_filename)
        sheet = workbook.active  # Assuming you're working with the active sheet

        # Check if the lead exists based on email
        lead_row = None
        for i, row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True)
        ):  # Start from the second row (assuming header)
            if row and row[0] == lead_email:  # Assuming email is in the first column
                lead_row = (
                    i + 2
                )  # Get the actual row number (adjust if email is in a different column)
                break

        if lead_row is None:
            logging.error(
                f"Error: Lead with email '{lead_email}' not found in '{excel_filename}'."
            )
            return False

        # Check if the lead_data fields match the column headers
        headers = [
            cell.value for cell in sheet[1]
        ]  # Get the column headers from the first row
        for key in lead_data.keys():
            if key not in headers:
                logging.error(f"Error: Column '{key}' not found in the Excel file.")
                return False

        # Add the lead data to the respective columns
        for key, value in lead_data.items():
            col_index = headers.index(key) + 1  # Get the column index (1-based)
            sheet.cell(row=lead_row, column=col_index, value=value)

        workbook.save(excel_filename)
        return True

    except FileNotFoundError:
        logging.error(f"Error: Excel file '{excel_filename}' not found.")
        return False
    except ValueError as e:
        logging.error(f"Error: {e}")
        return False
    except Exception as e:
        logging.error(f"An unexpected error occurred: {e}")
        return False


def preprocess_csv(raw_csv_file, columns_to_preprocess):
    """
    Preprocesses a CSV file by removing everything after a comma in specified columns.

    Args:
        raw_csv_file (str): The path to the CSV file.
        columns_to_preprocess (list): A list of column names to preprocess.
    """
    try:
        # Read the CSV file into a DataFrame
        df = pd.read_csv(raw_csv_file)

        # Preprocess the specified columns
        for column in columns_to_preprocess:
            if column in df.columns:
                df[column] = df[column].apply(
                    lambda x: x.split(",")[0] if isinstance(x, str) and "," in x else x
                )
            else:
                print(f"Warning: Column '{column}' not found in the CSV file.")

        # Overwrite the CSV file
        df.to_csv(raw_csv_file, index=False)
        print("CSV file preprocessed successfully.")

    except FileNotFoundError:
        print(f"Error: CSV file '{raw_csv_file}' not found.")
    except Exception as e:
        print(f"An error occurred during preprocessing: {e}")


def extract_lead_info_from_csv(raw_csv_file, lead_directory_file, column_mapping):
    """
    Extracts lead information from a raw CSV file and populates a lead directory Excel file.

    Args:
        raw_csv_file (str): Path to the raw CSV file.
        lead_directory_file (str): Path to the lead directory Excel file.
        column_mapping (dict): A dictionary mapping lead directory column names to
                               raw data columns. If a key is mapped to a list of columns,
                               the function will check each column in the list sequentially
                               and use the first non-empty value found.
                               Example: {"first_name": "First Name", "last_name": "Last Name",
                                        "email": ["Recommended Email", "Recommended Work Email", "Recommended Personal Email"],
                                        "company": "Employer"}

    Returns:
        bool: True if the lead directory was populated successfully, False otherwise.
    """
    try:
        # Load the raw CSV file
        with open(
            raw_csv_file, "r", newline="", encoding="utf-8"
        ) as csvfile:  # Adjust encoding if necessary
            reader = csv.DictReader(csvfile)
            raw_headers = reader.fieldnames

            # Load or create the lead directory Excel file
            try:
                lead_workbook = load_workbook(lead_directory_file)
            except FileNotFoundError:
                lead_workbook = (
                    Workbook()
                )  # Create a new workbook if the file doesn't exist

            lead_sheet = lead_workbook.active

            # Ensure the lead directory has the correct headers
            lead_headers = (
                [cell.value for cell in lead_sheet[1]] if lead_sheet.max_row > 0 else []
            )
            for lead_key in column_mapping.keys():
                if lead_key not in lead_headers:
                    lead_sheet.cell(row=1, column=len(lead_headers) + 1, value=lead_key)
                    lead_headers.append(lead_key)

            # Extract lead data and populate the lead directory
            for row in reader:
                lead_info = {}
                for lead_key, raw_columns in column_mapping.items():
                    if isinstance(raw_columns, str):
                        if raw_columns not in raw_headers:
                            logging.error(
                                f"Error: Column '{raw_columns}' not found in the raw data file."
                            )
                            return False
                        lead_info[lead_key] = row[
                            raw_columns
                        ].strip()  # Strip leading/trailing spaces
                    elif isinstance(raw_columns, list):
                        lead_info[lead_key] = None  # Initialize as None
                        for raw_column in raw_columns:
                            if raw_column not in raw_headers:
                                logging.error(
                                    f"Error: Column '{raw_column}' not found in the raw data file."
                                )
                                return False
                            email_value = row[
                                raw_column
                            ].strip()  # Strip leading/trailing spaces
                            if email_value and email_value != "-":
                                lead_info[lead_key] = email_value
                                break  # Stop checking after finding a valid value

                # Add the lead to the lead directory
                lead_sheet.append(
                    [lead_info.get(header, None) for header in lead_headers]
                )

            lead_workbook.save(lead_directory_file)
            return True

    except FileNotFoundError:
        logging.error(f"Error: Raw data file '{raw_csv_file}' not found.")
        return False
    except ValueError as e:
        logging.error(f"Error: {e}")
        return False
    except Exception as e:
        logging.error(f"An unexpected error occurred while processing the files: {e}")
        return False


def eliminate_incomplete_leads(excel_file, sheet_name, columns_to_check):
    """
    Eliminates leads from an Excel file if any of the specified key fields are missing.

    Args:
        excel_file (str): Path to the Excel file.
        sheet_name (str, optional): Name of the sheet to process. Defaults to "Sheet1".
        columns_to_check (list, optional): List of column names to check for missing values.
                                           Defaults to ["First Name", "Last Name", "Company Name", "Email"].

    Returns:
        None. Modifies the Excel file in place.
    """
    try:
        workbook = load_workbook(excel_file)
        sheet = workbook[sheet_name]

        # Get the header row and indices of key columns
        headers = [cell.value for cell in sheet[1]]
        key_indices = [headers.index(col) for col in columns_to_check if col in headers]

        # Identify and eliminate rows with missing key values
        eliminated_leads = []
        rows_to_delete = []
        for row_index, row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            lead_info = dict(zip(headers, row))
            if any(
                lead_info[key_col] is None or str(lead_info[key_col]).strip() == ""
                for key_col in columns_to_check
            ):
                eliminated_leads.append(lead_info)
                rows_to_delete.append(row_index)

        # Delete rows in reverse order to avoid index shifting
        for row_index in sorted(rows_to_delete, reverse=True):
            sheet.delete_rows(row_index)

        # Print eliminated leads
        if eliminated_leads:
            print("Eliminated Leads:")
            print("| " + " | ".join(headers) + " |")
            print("| " + " | ".join("-" * len(header) for header in headers) + " |")
            for lead in eliminated_leads:
                print(
                    "| "
                    + " | ".join(str(lead.get(header, "")) for header in headers)
                    + " |"
                )

        workbook.save(excel_file)

    except FileNotFoundError:
        print(f"Error: Excel file '{excel_file}' not found.")
    except ValueError as e:
        print(f"Error: {e}")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")


def clean_emails(excel_file, sheet_name, columns_to_clean, row_start, row_end):
    """
    Cleans the emails in the specified columns of an Excel sheet by removing everything before the <html> tag and after the </html> tag.

    Args:
        excel_file (str): The path to the Excel file.
        sheet_name (str): The name of the sheet containing the emails to be cleaned.
        columns_to_clean (list): A list of column names containing the emails to be cleaned.
        row_start (int): The starting row number (inclusive) for cleaning emails.
        row_end (int): The ending row number (inclusive) for cleaning emails.
    """
    try:
        workbook = load_workbook(excel_file)
        sheet = workbook[sheet_name]
        headers = [cell.value for cell in sheet[1]]

        for column_name in columns_to_clean:
            if column_name not in headers:
                print(f"Error: Column '{column_name}' not found in the Excel file.")
                continue

            col_index = headers.index(column_name) + 1
            for row_index in range(row_start, row_end + 1):
                cell = sheet.cell(row=row_index, column=col_index)
                if cell.value is not None:
                    email = str(cell.value)
                    # Remove everything before <html> and after </html>
                    email = email.split("<html>")[1].split("</html>")[0]

                    cell.value = (
                        "<html>" + email + "</html>"
                    )  # Add back the <html> and </html> tags

        workbook.save(excel_file)  # Save the changes to the same Excel file
        print("Emails cleaned successfully.")

    except FileNotFoundError:
        print(f"Error: Excel file '{excel_file}' not found.")
    except KeyError:
        print(f"Error: Sheet '{sheet_name}' not found in the Excel file.")
    except Exception as e:
        print(f"An error occurred: {e}")


def add_opt_out_message(excel_file, sheet_name, columns_to_update, row_start, row_end):
    """
    Adds an opt-out message to the footer of HTML emails in specified columns and rows of an Excel sheet.

    Args:
        excel_file (str): The path to the Excel file.
        sheet_name (str): The name of the sheet containing the emails.
        columns_to_update (list): A list of column names containing the emails to be updated.
        row_start (int): The starting row number (inclusive) for adding the opt-out message.
        row_end (int): The ending row number (inclusive) for adding the opt-out message.
    """
    try:
        workbook = load_workbook(excel_file)
        sheet = workbook[sheet_name]
        headers = [cell.value for cell in sheet[1]]

        opt_out_message = """
        <footer>
          <p style="font-size: small; color: #666666;">To opt out of further emails, please reply with "Stop".</p>
        </footer>
        """

        for column_name in columns_to_update:
            if column_name not in headers:
                print(f"Error: Column '{column_name}' not found in the Excel file.")
                continue

            col_index = headers.index(column_name) + 1
            for row_index in range(row_start, row_end + 1):
                cell = sheet.cell(row=row_index, column=col_index)
                if (
                    cell.value is not None
                    and isinstance(cell.value, str)
                    and "<html>" in cell.value
                    and "</html>" in cell.value
                ):
                    email = cell.value
                    email = email.replace("</body>", f"{opt_out_message}</body>")
                    cell.value = email

        workbook.save(excel_file)
        print("Opt-out messages added successfully.")

    except FileNotFoundError:
        print(f"Error: Excel file '{excel_file}' not found.")
    except KeyError:
        print(f"Error: Sheet '{sheet_name}' not found in the Excel file.")
    except Exception as e:
        print(f"An error occurred: {e}")


def email_verifier(csv_file, excel_file, mapping, output_file):
    """
    Verifies emails in an Excel file against a cleaned CSV file and creates a new Excel file with only verified emails.

    Args:
        csv_file (str): Path to the CSV file with cleaned emails.
        excel_file (str): Path to the Excel file with original data.
        mapping (dict): A dictionary mapping Excel email column to CSV email column.
        Use excel_email_column and csv_email_column keys.
        output_file (str): Path to save the output Excel file.
    """
    try:
        # 1. Check Column Existence and Load Data
        logging.info(
            f"Starting verification process. CSV: {csv_file}, Excel: {excel_file}"
        )

        try:
            df_csv = pd.read_csv(csv_file)
        except FileNotFoundError:
            logging.error(f"CSV file not found: {csv_file}")
            return

        try:
            df_excel = pd.read_excel(excel_file)
        except FileNotFoundError:
            logging.error(f"Excel file not found: {excel_file}")
            return

        excel_email_col = mapping.get("excel_email_column")
        csv_email_col = mapping.get("csv_email_column")

        if not excel_email_col or excel_email_col not in df_excel.columns:
            logging.error(f"Email column '{excel_email_col}' not found in Excel file.")
            return

        if not csv_email_col or csv_email_col not in df_csv.columns:
            logging.error(f"Email column '{csv_email_col}' not found in CSV file.")
            return

        # 2. Iterate and Store Rows (Using sets for efficiency)
        rows_to_keep = set()
        for email in df_csv[csv_email_col]:
            # Find matching rows in Excel (case-insensitive and stripped of whitespace)
            matching_rows = df_excel.index[
                df_excel[excel_email_col].str.lower().str.strip()
                == email.lower().strip()
            ].tolist()
            rows_to_keep.update(matching_rows)

        # 3. Keep only verified rows
        df_verified = df_excel.loc[list(rows_to_keep)]

        # 4. Output to New Excel File
        df_verified.to_excel(output_file, index=False)
        logging.info(f"Verification complete. Output saved to: {output_file}")

    except Exception as e:
        logging.exception(f"An error occurred during the verification process: {e}")


def transfer_data_between_files(excel_file, csv_file, mapping_dict):
    """
    Transfers data from a CSV file to an Excel file based on a mapping dictionary.

    Args:
        excel_file: Path to the Excel file.
        csv_file: Path to the raw CSV file.
        mapping_dict: A dictionary with two elements:
            - 'keys': A dictionary mapping the Excel key column name to the CSV key column name.
            - 'data': A dictionary mapping the CSV data column name to the desired new column name in Excel.
    """

    # --- 1. Load the data ---
    try:
        df_excel = pd.read_excel(excel_file)
    except FileNotFoundError:
        print(f"Error: Excel file not found at {excel_file}")
        return
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return

    try:
        df_csv = pd.read_csv(csv_file)
    except FileNotFoundError:
        print(f"Error: CSV file not found at {csv_file}")
        return
    except Exception as e:
        print(f"Error reading CSV file: {e}")
        return

    # --- 2. Validate the mapping dictionary ---
    if (
        not isinstance(mapping_dict, dict)
        or "keys" not in mapping_dict
        or "data" not in mapping_dict
    ):
        print(
            "Error: Invalid mapping_dict. It should be a dictionary with 'keys' and 'data'."
        )
        return

    excel_key_col = list(mapping_dict["keys"].keys())[0]
    csv_key_col = mapping_dict["keys"][excel_key_col]
    csv_data_col = list(mapping_dict["data"].keys())[0]
    new_excel_col = mapping_dict["data"][csv_data_col]

    if not (
        excel_key_col in df_excel.columns
        and csv_key_col in df_csv.columns
        and csv_data_col in df_csv.columns
    ):
        print(f"Error: One or more specified columns not found in the files.")
        print(f"  Excel key column: {excel_key_col}")
        print(f"  CSV key column: {csv_key_col}")
        print(f"  CSV data column: {csv_data_col}")
        return

    # --- 3. Create a data map from CSV (case-insensitive for keys) ---
    data_map = {
        str(row[csv_key_col]).lower(): row[csv_data_col]
        for index, row in df_csv.iterrows()
    }

    # --- 4. Add the new column to the Excel dataframe if it doesn't exist ---
    if new_excel_col not in df_excel.columns:
        df_excel[new_excel_col] = ""  # Initialize with empty strings

    # --- 5. Populate the new column with data from the map ---
    for index, row in df_excel.iterrows():
        key = str(row[excel_key_col]).lower()
        data_value = data_map.get(key, "")  # Get data or "" if not found
        # Only update if cell is empty or if we're overwriting with non-empty data
        if pd.isna(df_excel.at[index, new_excel_col]) or data_value != "":
            df_excel.at[index, new_excel_col] = data_value

    # --- 6. Save the updated Excel file ---
    try:
        df_excel.to_excel(excel_file, index=False)
        print(f"Successfully added/updated '{new_excel_col}' column in {excel_file}")
    except Exception as e:
        print(f"Error saving updated Excel file: {e}")
